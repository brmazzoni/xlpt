from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

from styles import *


def gen_template(headers, header_line=1, header_fill_color=HEADER_FILL_COLOR, filename='template.xlsx', metadata=None):

  # Define (default medium gray) solid color fill for headers
  header_fill = PatternFill(
    fill_type='solid', 
    start_color=header_fill_color,
    end_color=header_fill_color)

  # Initiate new workbook and select worksheet
  wb = Workbook()
  ws = wb.active

  # Write metadata
  if metadata is not None:
    for cell, content in metadata.items():
      ws[cell] = content['text']


  # Generate headers
  for col, header in headers.items():
    cell = col + str(header_line)
    ws[cell] = header['text']
    ws[cell].fill = header_fill
    width = header['width'] if 'width' in header.keys() else CW_NORMAL
    ws.column_dimensions[col].width = width
  

  # Style columns (10 empty lines)
  EMPTY_LINES = 10
  for row in ws[f'A{header_line}:{max(headers)}{EMPTY_LINES+header_line}']:
    for cell in row:
      cell.border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000')
      )
    
  for col, header in headers.items():
    for cell in ws[f'{col}{header_line}:{col}{EMPTY_LINES+header_line}']:
      wrap_text = header['wrap'] if 'wrap' in header.keys() else False
      cell[0].alignment = Alignment(wrap_text=wrap_text)


  # Save file
  wb.save(filename)

if __name__ == '__main__':

  metadata = {
    'B1': {'text': 'Title'},  'C1': {'text': 'version'},  'D1': {'text': '1.0'},
                              'C2': {'text': 'author'},   'D2': {'text': 'BMA'},
    }

  headers = {
    'A': {'text': 'c1', 'width': CW_SHORT},
    'B': {'text': 'column2', 'width': CW_LONG, 'wrap': True},
    'C': {'text': 'column3', 'wrap_text': True},
  }
  gen_template(headers, header_line=5, metadata=metadata)
