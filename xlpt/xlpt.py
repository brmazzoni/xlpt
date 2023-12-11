import sys

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from styles import *

B_THIN = Side(border_style='thin', color='FF000000')

def all_borders(ws_slice):
  for row in ws_slice:
    for cell in row:
      cell.border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000')
      )

def outer_borders(ws_slice):
  rows = ws_slice
  if len(rows) == 1:
    _outer_borders_single_row(ws, _range)
    return

  for i, row in enumerate(rows):
    for j, cell in enumerate(row):
      # Corners first
      if (i, j) == (0, 0):
        cell.border = Border(top=B_THIN, left=B_THIN)
      elif (i, j) == (len(rows)-1, 0):
        cell.border = Border(bottom=B_THIN, left=B_THIN)
      elif (i, j) == (0, len(row)-1):
        cell.border = Border(top=B_THIN, right=B_THIN)
      elif (i, j) == (len(rows)-1, len(row)-1):
        cell.border = Border(bottom=B_THIN, right=B_THIN)
      # then edges\corners
      elif i == 0:
        cell.border = Border(top=B_THIN)
      elif i == len(rows) - 1:
        cell.border = Border(bottom=B_THIN)
      elif j == 0:
        cell.border = Border(left=B_THIN)
      elif j == len(row) - 1:
        cell.border = Border(right=B_THIN)


def _outer_borders_single_row(row_cells):
  cells = row_cells
  for i, c in enumerate(cells):
    if i == 0:
      c.border = Border(top=B_THIN, bottom=B_THIN, left=B_THIN)
    elif i == len(cells) - 1:
      c.border = Border(top=B_THIN, bottom=B_THIN, right=B_THIN)
    else:
      c.border = Border(top=B_THIN, bottom=B_THIN)

def all_horizontal_borders(ws_slice):
  for row in ws_slice:
    _outer_borders_single_row(row)

def draw_borders(ws_slice, borders='all'):
  if borders == 'all':
    all_borders(ws_slice)
  elif borders == 'outer':
    outer_borders(ws_slice)
  elif borders == 'all_horizontal':
    all_horizontal_borders(ws_slice)
  else:
    raise(Exception(f'Error: unknown value for borders style: {borders}'))


def build(headers, header_line=1, header_fill_color=HEADER_FILL_COLOR, filename='template.xlsx', metadata=None):

  # Define (default medium gray) solid color fill for headers
  header_fill = PatternFill(
    fill_type='solid', 
    start_color=header_fill_color,
    end_color=header_fill_color)

  # Initiate new workbook and select worksheet
  wb = Workbook()
  ws = wb.active

  all_horizontal_borders(ws['H3:J6'])
  
  # Write metadata
  if metadata is not None:
    for section, data in metadata.items():
      if 'borders' in data.keys():
        draw_borders(ws[data['range']], data['borders'])
      if 'text' in data.keys():
        for cell, text in data['text'].items():
          ws[cell] = text
          ws[cell].font = Font(size=data['font'])
      for row in ws[data['range']]:
        for cell in row:
          pass
          
          #ws[f'{cell.column_letter}{cell.row}'] = section
        

  # Generate headers
  for col, header in headers.items():
    cell = col + str(header_line)
    ws[cell] = header['text']
    ws[cell].fill = header_fill
    width = header['width'] if 'width' in header.keys() else CW_NORMAL
    ws.column_dimensions[col].width = width
  

  # Style columns (10 empty lines)
  EMPTY_LINES = 10
  for row in ws[f'A{header_line}:{max(headers)}{EMPTY_LINES+header_line}']:
    for cell in row:
      cell.border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000')
      )
    
  for col, header in headers.items():
    for cell in ws[f'{col}{header_line}:{col}{EMPTY_LINES+header_line}']:
      wrap_text = header['wrap'] if 'wrap' in header.keys() else False
      cell[0].alignment = Alignment(wrap_text=wrap_text)


  # Save file
  wb.save(filename)


if __name__ == '__main__':
  metadata = {
    'section1': {
      'range': 'A1:B3', 
      'image': 'img/logo.png',
      'borders': 'outer' # outer/all
    },
    'section2': {
      'range': 'C1:D1',
      'text': {'C1': 'TITLE'},
      'font': 24
    },
    'section3': {
      'range': 'E1:F1',
      'text': {'E1': 'Subtitle'},
      'font': 12
    },
    'section4': {
      'range': 'C2:D3',
      'text': {'C2': 'author', 'C3': 'version'},
      'font': 10,
      'borders': 'outer'
    },
    'section5': {
      'range': 'E2:F3',
      'text': {'E2': '<author name>', 'E3': '<version number>'},
      'font': 10,
      'borders': 'all_horizontal'
    }
  }

  headers = {
    'A': {'text': 'c1', 'width': CW_SHORT},
    'B': {'text': 'column2', 'width': CW_LONG, 'wrap': True},
    'C': {'text': 'column3', 'wrap_text': True},
  }
    
  build(headers, metadata=metadata, header_line=5)
